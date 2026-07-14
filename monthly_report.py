#!/usr/bin/env python3
"""
monthly_report.py — 8-combination monthly performance report
=============================================================
Entry:    3:15pm (fixed) across all combinations
Params:   4 sets — baseline (LB=30, VM=7) + top-3 from sweep (LB=36/37/35, VM=6)
Exits:    Standard (100% at next-day 3pm open) | Split (50% at 9:45 + 50% at 11am)
Sizing:   ₹6L pool — 1-6 signals → ₹1L each; 7+ → ₹6L ÷ n (floor shares, no compounding)

Output:   results/monthly_performance_report.xlsx  (8 sheets)
"""

import sys
import warnings
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE))

from prepare_data import build_diagnostic_table

RESULTS_DIR  = BASE / "results"
ENTRY_COL    = "entry_price_315pm"
POOL         = 600_000   # ₹6,00,000
MAX_STOCK    = 100_000   # ₹1,00,000 per stock
THRESHOLD    = 6         # n ≤ 6 → ₹1L each; n ≥ 7 → pool/n
START_EQUITY = 600_000   # equity curve starts at ₹6L

PARAM_SETS = [
    {"label": "Baseline LB=30 VM=7",  "lookback": 30, "vol_mult": 7, "short": "LB30_VM7"},
    {"label": "Top-1  LB=36 VM=6",    "lookback": 36, "vol_mult": 6, "short": "LB36_VM6"},
    {"label": "Top-2  LB=37 VM=6",    "lookback": 37, "vol_mult": 6, "short": "LB37_VM6"},
    {"label": "Top-3  LB=35 VM=6",    "lookback": 35, "vol_mult": 6, "short": "LB35_VM6"},
]

EXIT_STRATEGIES = [
    {
        "label":     "Standard (3pm)",
        "short":     "Std",
        "exit_type": "single",
        "exit_cols": "exit_3pm_open",
    },
    {
        "label":     "Split (9:45+11am)",
        "short":     "Split",
        "exit_type": "split",
        "exit_cols": ("exit_945_open", "exit_1100_open"),
    },
]


# ── Backtest with custom position sizing ──────────────────────────────────────

def run_backtest(signals, exit_type, exit_cols):
    """
    Run backtest for 3:15pm entry with ₹6L pool sizing.

    Position sizing:
      n ≤ 6  →  ₹1,00,000 each
      n ≥ 7  →  ₹6,00,000 ÷ n each
    Whole shares (floor division), no compounding.
    """
    if exit_type == "split":
        req_cols = list(exit_cols)
    else:
        req_cols = [exit_cols]

    valid = signals.dropna(subset=[ENTRY_COL] + req_cols).copy()

    by_date = defaultdict(list)
    for _, row in valid.iterrows():
        by_date[row["date"]].append(row)

    exec_rows = []
    for entry_date in sorted(by_date):
        day_sigs = by_date[entry_date]
        n        = len(day_sigs)
        target   = MAX_STOCK if n <= THRESHOLD else POOL / n

        for row in day_sigs:
            ep     = float(row[ENTRY_COL])
            shares = int(target // ep)
            if shares == 0:
                continue

            actual_cap = shares * ep

            if exit_type == "split":
                col_945, col_1100 = exit_cols
                x945         = float(row[col_945])
                x1100        = float(row[col_1100])
                sh_first     = shares // 2
                sh_second    = shares - sh_first
                pnl          = sh_first * (x945 - ep) + sh_second * (x1100 - ep)
                blended      = (x945 + x1100) / 2
                ret_pct      = (blended - ep) / ep * 100
            else:
                xp      = float(row[exit_cols])
                pnl     = shares * (xp - ep)
                ret_pct = (xp - ep) / ep * 100

            exec_rows.append({
                "date":              entry_date,
                "symbol":            row["symbol"],
                "entry_price":       round(ep, 4),
                "shares":            shares,
                "capital_allocated": round(actual_cap, 2),
                "pnl":               round(pnl, 2),
                "return_pct":        round(ret_pct, 4),
                "n_signals_day":     n,
            })

    if not exec_rows:
        return pd.DataFrame()

    trades = pd.DataFrame(exec_rows)
    trades["date"] = pd.to_datetime(trades["date"])
    return trades.sort_values(["date", "symbol"]).reset_index(drop=True)


# ── Equity curve helpers ──────────────────────────────────────────────────────

def build_equity_curve(trades):
    """
    Build a day-indexed Series of equity values, running peak, drawdown (₹ and %).
    Returns a DataFrame indexed by trading date.
    """
    daily_pnl = (trades.groupby("date")["pnl"]
                 .sum()
                 .sort_index())
    ec = pd.DataFrame({"pnl": daily_pnl})
    ec["equity"]       = START_EQUITY + ec["pnl"].cumsum()
    ec["running_peak"] = ec["equity"].cummax()
    ec["dd_rs"]        = ec["running_peak"] - ec["equity"]
    ec["dd_pct"]       = ec["dd_rs"] / ec["running_peak"] * 100
    return ec


# ── Monthly breakdown ─────────────────────────────────────────────────────────

def compute_monthly_table(trades):
    """Return a DataFrame with one row per calendar month.

    Months with no trades are included as zero rows so there are no gaps
    in the output table.
    """
    if trades.empty:
        return pd.DataFrame()

    trades = trades.copy()
    trades["date"]       = pd.to_datetime(trades["date"])
    trades["year_month"] = trades["date"].dt.to_period("M")

    ec = build_equity_curve(trades)

    # Full month range from first to last trade month (no gaps)
    first_ym = trades["year_month"].min()
    last_ym  = trades["year_month"].max()
    all_months = pd.period_range(first_ym, last_ym, freq="M")

    rows = []
    for ym in all_months:
        m_trades = trades[trades["year_month"] == ym]
        m_dates  = m_trades["date"].unique()
        m_ec     = ec.loc[ec.index.isin(m_dates)]

        if len(m_trades) == 0:
            # No trades this month — show zero row
            rows.append({
                "Month":                 str(ym),
                "Revenue (₹)":           0.0,
                "Biggest 1-Day Loss (₹)": 0.0,
                "Max Drawdown (₹)":      0.0,
                "Max Drawdown (%)":      0.0,
                "Trades":                0,
                "Sessions":              0,
                "Win %":                 "",
                "Loss %":                "",
                "Avg Win Return %":      "",
                "Avg Loss Return %":     "",
                "Avg Return / Trade %":  "",
            })
            continue

        # 1 — Revenue (₹)
        revenue = m_trades["pnl"].sum()

        # 2 — Biggest single-day loss (₹)  [most negative daily total]
        daily_sums   = m_trades.groupby("date")["pnl"].sum()
        biggest_loss = daily_sums.min()

        # 3 — Max drawdown within this month (continuous curve)
        if not m_ec.empty:
            idx_max    = m_ec["dd_rs"].idxmax()
            max_dd_rs  = m_ec.loc[idx_max, "dd_rs"]
            max_dd_pct = m_ec.loc[idx_max, "dd_pct"]
        else:
            max_dd_rs = max_dd_pct = 0.0

        # 4 — Trades count
        n_trades = len(m_trades)

        # 5 — Trading sessions (unique entry days)
        n_sessions = m_trades["date"].nunique()

        # 6 — Win / loss %
        n_pos   = (m_trades["pnl"] > 0).sum()
        n_neg   = n_trades - n_pos
        pct_pos = n_pos / n_trades * 100
        pct_neg = n_neg / n_trades * 100

        # 7 — Average return for winners and losers separately
        winners  = m_trades.loc[m_trades["pnl"] > 0,  "return_pct"]
        losers   = m_trades.loc[m_trades["pnl"] <= 0, "return_pct"]
        avg_win  = winners.mean() if len(winners) > 0 else np.nan
        avg_loss = losers.mean()  if len(losers)  > 0 else np.nan

        # 8 — Average return per trade (all trades)
        avg_all = m_trades["return_pct"].mean()

        rows.append({
            "Month":                     str(ym),
            "Revenue (₹)":               round(revenue, 2),
            "Biggest 1-Day Loss (₹)":    round(biggest_loss, 2),
            "Max Drawdown (₹)":          round(max_dd_rs, 2),
            "Max Drawdown (%)":          round(max_dd_pct, 4),
            "Trades":                    n_trades,
            "Sessions":                  n_sessions,
            "Win %":                     round(pct_pos, 2),
            "Loss %":                    round(pct_neg, 2),
            "Avg Win Return %":          round(avg_win,  4) if pd.notna(avg_win)  else "",
            "Avg Loss Return %":         round(avg_loss, 4) if pd.notna(avg_loss) else "",
            "Avg Return / Trade %":      round(avg_all,  4),
        })

    return pd.DataFrame(rows)


# ── Summary stats ─────────────────────────────────────────────────────────────

def compute_summary(trades, combo_label):
    if trades.empty:
        return {
            "Combination":               combo_label,
            "Total Revenue (₹)":         0,
            "Overall Max DD (₹)":        0,
            "Overall Max DD (%)":        0,
            "Biggest 1-Day Loss (₹)":    0,
            "Overall Win Rate %":        0,
            "Avg Return / Trade %":      0,
            "Total Trades":              0,
        }

    ec        = build_equity_curve(trades)
    idx_max   = ec["dd_rs"].idxmax()

    return {
        "Combination":            combo_label,
        "Total Revenue (₹)":      round(trades["pnl"].sum(), 2),
        "Overall Max DD (₹)":     round(ec.loc[idx_max, "dd_rs"],  2),
        "Overall Max DD (%)":     round(ec.loc[idx_max, "dd_pct"], 4),
        "Biggest 1-Day Loss (₹)": round(trades.groupby("date")["pnl"].sum().min(), 2),
        "Overall Win Rate %":     round((trades["pnl"] > 0).mean() * 100, 2),
        "Avg Return / Trade %":   round(trades["return_pct"].mean(), 4),
        "Total Trades":           len(trades),
    }


# ── Excel formatting helpers ──────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT   = Font(color="FFFFFF", bold=True, size=11)
NORMAL_FONT  = Font(size=9)
BOLD_FONT    = Font(bold=True, size=9)

COL_WIDTHS = {
    "Month":                  12,
    "Revenue (₹)":            16,
    "Biggest 1-Day Loss (₹)": 22,
    "Max Drawdown (₹)":       18,
    "Max Drawdown (%)":       18,
    "Trades":                 9,
    "Sessions":               11,
    "Win %":                  9,
    "Loss %":                 9,
    "Avg Win Return %":       18,
    "Avg Loss Return %":      19,
    "Avg Return / Trade %":   21,
}

RS_FMT  = '#,##0.00'
PCT_FMT = '0.00'
INT_FMT = '0'
NM4_FMT = '0.0000'

COL_FORMATS = {
    "Revenue (₹)":            RS_FMT,
    "Biggest 1-Day Loss (₹)": RS_FMT,
    "Max Drawdown (₹)":       RS_FMT,
    "Max Drawdown (%)":       PCT_FMT,
    "Trades":                 INT_FMT,
    "Sessions":               INT_FMT,
    "Win %":                  PCT_FMT,
    "Loss %":                 PCT_FMT,
    "Avg Win Return %":       NM4_FMT,
    "Avg Loss Return %":      NM4_FMT,
    "Avg Return / Trade %":   NM4_FMT,
}


def format_sheet(ws, monthly_df, combo_label):
    cols = list(monthly_df.columns)
    n_cols = len(cols)

    # Row 1: title
    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value=combo_label)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Row 2: blank spacer
    ws.row_dimensions[2].height = 4

    # Row 3: headers
    ws.row_dimensions[3].height = 18
    for ci, col_name in enumerate(cols, start=1):
        c = ws.cell(row=3, column=ci, value=col_name)
        c.font   = HEADER_FONT
        c.fill   = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 4+: data
    for ri, row_data in enumerate(monthly_df.itertuples(index=False), start=4):
        fill = ALT_FILL if (ri % 2 == 0) else None
        ws.row_dimensions[ri].height = 15
        for ci, col_name in enumerate(cols, start=1):
            val = row_data[ci - 1]
            c = ws.cell(row=ri, column=ci, value=(None if val == "" else val))
            c.font = NORMAL_FONT
            c.alignment = Alignment(horizontal="right" if ci > 1 else "center",
                                    vertical="center")
            if fill:
                c.fill = fill
            fmt = COL_FORMATS.get(col_name)
            if fmt and val != "" and val is not None:
                c.number_format = fmt

    # Column widths
    for ci, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 14)

    # Freeze panes (keep header visible)
    ws.freeze_panes = "B4"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    RESULTS_DIR.mkdir(exist_ok=True)
    output_path = RESULTS_DIR / "monthly_performance_report.xlsx"

    print("=" * 72)
    print("monthly_report.py — 8-combination monthly performance report")
    print(f"Entry:  3:15pm  |  Pool: ₹6L  |  Threshold: ≤{THRESHOLD} stocks → ₹1L each")
    print(f"Output: {output_path}")
    print("=" * 72)

    # ── Build one diagnostic table per unique lookback ────────────────────────
    unique_lookbacks = sorted({ps["lookback"] for ps in PARAM_SETS})
    diag_cache = {}

    for lb in unique_lookbacks:
        print(f"\n[1/2] Building diagnostic table: lookback={lb}d …", flush=True)
        diag = build_diagnostic_table(
            vol_window=lb,
            vol_mult=0.0,   # vol_mult=0 → keep all rows; filter by vol_mult below
            save_csv=False,
            verbose=False,
        )
        diag["date"] = pd.to_datetime(diag["date"])
        diag_cache[lb] = diag
        print(f"      {len(diag):,} rows  |  {int(diag['passes_return'].sum()):,} pass_return")

    # ── Run 8 combinations ────────────────────────────────────────────────────
    print("\n[2/2] Running 8 combinations …")
    sheets_data  = {}   # sheet_name → (combo_label, monthly_df)
    all_summaries = []

    combo_order = []   # preserve sheet order
    for ps in PARAM_SETS:
        for ex in EXIT_STRATEGIES:
            lb      = ps["lookback"]
            mult    = ps["vol_mult"]
            diag    = diag_cache[lb]
            signals = diag[(diag["volume_ratio"] >= mult) & diag["passes_return"]].copy()

            combo_label = f"{ps['label']}  |  {ex['label']}"
            sheet_name  = f"{ps['short']}_{ex['short']}"

            print(f"  {combo_label} …", end=" ", flush=True)
            trades = run_backtest(signals, ex["exit_type"], ex["exit_cols"])
            print(f"{len(trades):,} trades")

            monthly_df = compute_monthly_table(trades)
            summary    = compute_summary(trades, combo_label)

            sheets_data[sheet_name] = (combo_label, monthly_df)
            combo_order.append(sheet_name)
            all_summaries.append(summary)

    # ── Write Excel ───────────────────────────────────────────────────────────
    print(f"\nWriting {output_path} …")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name in combo_order:
            combo_label, monthly_df = sheets_data[sheet_name]
            if monthly_df.empty:
                monthly_df = pd.DataFrame(columns=[
                    "Month","Revenue (₹)","Biggest 1-Day Loss (₹)",
                    "Max Drawdown (₹)","Max Drawdown (%)","Trades","Sessions",
                    "Win %","Loss %","Avg Win Return %","Avg Loss Return %",
                    "Avg Return / Trade %",
                ])
            # Write data starting at row 3 so we can add title in row 1
            monthly_df.to_excel(writer, sheet_name=sheet_name,
                                 index=False, startrow=2)
            ws = writer.sheets[sheet_name]
            format_sheet(ws, monthly_df, combo_label)

    print(f"Saved: {output_path}")

    # ── Console summary ───────────────────────────────────────────────────────
    summary_df = pd.DataFrame(all_summaries)

    print("\n" + "=" * 110)
    print("SUMMARY — 8 Combinations")
    print("=" * 110)
    print(f"{'#':<2}  {'Combination':<45}  {'Revenue (₹)':>14}  {'Max DD (₹)':>12}  "
          f"{'Max DD %':>9}  {'Worst Day (₹)':>14}  {'Win%':>6}  {'Avg%':>8}  {'Trades':>7}")
    print("    " + "─" * 104)
    for i, row in summary_df.iterrows():
        print(
            f"{i+1:<2}  {row['Combination']:<45}  "
            f"{row['Total Revenue (₹)']:>14,.0f}  "
            f"{row['Overall Max DD (₹)']:>12,.0f}  "
            f"{row['Overall Max DD (%)']:>8.2f}%  "
            f"{row['Biggest 1-Day Loss (₹)']:>14,.0f}  "
            f"{row['Overall Win Rate %']:>5.1f}%  "
            f"{row['Avg Return / Trade %']:>+8.4f}  "
            f"{int(row['Total Trades']):>7,}"
        )
    print("=" * 110)


if __name__ == "__main__":
    main()
